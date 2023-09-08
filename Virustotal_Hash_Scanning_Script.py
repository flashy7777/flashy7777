#This script will scan all hashes from an excel file and present output in another excel file which will contain hash detection report by several vendors through virustotal api.
#In order to run the script successfully packages such as vt, openpyxl must be installed in editor environment.
import vt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Enter your API key
client = vt.Client("73c23328fecf569ae22e349029e23dd7f524ad8b7749a2c611724ef466d8b23a")
#5cdec7ac061fdb5f9a918bd9195d732290d2547e254ca334627772f490798739
# Load file hashes from an Excel file
workbook = openpyxl.load_workbook("file_hashes.xlsx")
sheet = workbook.active
file_hashes = []
for row in sheet.iter_rows(min_row=1, values_only=True):
    hash_value = row[0]
    if hash_value:  # skip empty cells
        file_hashes.append(hash_value)
# List of vendors to check for detections
vendors = ["Symantec", "TrendMicro", "McAfee", "Sophos", "Microsoft", "SentinelOne", "CrowdStrike"]

# Create a new workbook and worksheet to store the results
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = "Scan Results"
output_sheet.cell(row=1, column=1, value="Hashes")
for i, vendor in enumerate(vendors):
    output_sheet.cell(row=1, column=i+2, value=vendor)

# Define cell styles
yes_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yes_font = Font(name='Calibri', size=11, color='006100')
yes_alignment = Alignment(horizontal='center', vertical='center')
no_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
no_font = Font(name='Calibri', size=11, color='9C0006')
no_alignment = Alignment(horizontal='center', vertical='center')

# Set header cell styles
output_sheet.cell(row=1, column=1).font = Font(name='Calibri', size=11, bold=True)
for col, vendor in enumerate(vendors, start=4):
    output_sheet.cell(row=1, column=col, value=vendor).font = Font(name='Calibri', size=11, bold=True)

# Loop through each hash and check the detections for each vendor
for row, hash in enumerate(file_hashes, start=2):
    output_sheet.cell(row=row, column=1, value=hash).font = Font(name='Calibri', size=11)
    try:
        analysis = client.get_object("/files/{}".format(hash), params={"allinfo": 1})
        output_sheet.cell(row=row, column=2, value=analysis.md5).font = Font(name='Calibri', size=11)
        output_sheet.cell(row=row, column=3, value=analysis.sha256).font = Font(name='Calibri', size=11)
        for col, vendor in enumerate(vendors, start=4):
            if vendor in analysis.last_analysis_results and analysis.last_analysis_results[vendor]["category"] == "malicious":
                output_sheet.cell(row=row, column=col, value="YES").fill = yes_fill
                output_sheet.cell(row=row, column=col).font = yes_font
                output_sheet.cell(row=row, column=col).alignment = yes_alignment
            else:
                output_sheet.cell(row=row, column=col, value="NO").fill = no_fill
                output_sheet.cell(row=row, column=col).font = no_font
                output_sheet.cell(row=row, column=col).alignment = no_alignment
    except vt.APIError as e:
        print("Error retrieving analysis results for {}: {}".format(hash, e))

# Save the output workbook
output_workbook.save("scan_results.xlsx")
